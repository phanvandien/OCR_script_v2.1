import os
import zipfile
import io
import json
import time
import re
import base64
import pandas as pd
import logging
import datetime
import uuid
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from openpyxl.utils import get_column_letter
from .forms import UploadZipForm
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.views.decorators.csrf import csrf_exempt
from PIL import Image
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import redis

# Redis client
redis_client = redis.Redis.from_url(settings.CELERY_BROKER_URL)

# Setup logging
logger = logging.getLogger('apps')

# --- LLM Setup ---
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage

# --- Pydantic models ---
from pydantic import BaseModel, Field
from typing import List, Optional

class TranscriptItem(BaseModel):
    Sbd: str
    Thi: Optional[float] = None

class TranscriptData(BaseModel):
    items: List[TranscriptItem]

class CertificateItem(BaseModel):
    Bang_cap: str
    Nganh: str
    Noi_cap: str
    Ho_ten: str
    Date_birth_VN: str

class CertificateData(BaseModel):
    items: List[CertificateItem]

# Prompts
TRANSCRIPT_PROMPT = """
Trích xuất danh sách sinh viên từ ảnh bảng điểm.
Trích xuất: SBD (số báo danh), Thi (điểm thi)
Trả về JSON: {"items": [{"Sbd": "00123", "Thi": 8.5}]}
"""

CERTIFICATE_PROMPT = """
Trích xuất thông tin từ ảnh văn bằng/chứng chỉ (phần tiếng Việt).
Trả về JSON: {
  "items": [{
    "Bang_cap": "tên bằng cấp",
    "Nganh": "tên ngành học", 
    "Noi_cap": "tên trường",
    "Ho_ten": "họ tên",
    "Date_birth_VN": "dd/mm/yyyy"
  }]
}
"""

def update_progress(session_id, current, total, message=""):
    """Update progress in Redis"""
    try:
        percentage = (current / total * 100) if total > 0 else 0
        progress_data = {
            'current': current,
            'total': total,
            'percentage': round(percentage, 1),
            'message': message,
            'timestamp': time.time()
        }
        redis_client.set(f"progress:{session_id}", json.dumps(progress_data), ex=3600)
        logger.info(f"Progress: {session_id} - {current}/{total} ({percentage:.1f}%)")
    except Exception as e:
        logger.error(f"Error updating progress: {e}")

def get_progress(session_id):
    """Get progress from Redis"""
    try:
        data = redis_client.get(f"progress:{session_id}")
        if data:
            return json.loads(data)
    except Exception as e:
        logger.error(f"Error getting progress: {e}")
    
    return {'current': 0, 'total': 0, 'percentage': 0, 'message': 'Đang khởi tạo...'}

def clean_sbd(raw_value: str) -> str:
    """Clean and format SBD to 5 digits"""
    if not raw_value:
        return ""
    
    digits = re.findall(r"\d+", str(raw_value))
    if not digits:
        return str(raw_value)
    
    number_str = "".join(digits)
    return number_str.zfill(5)[:5]

def clean_date_string(date_str: str) -> str:
    """Clean date string to dd/mm/yyyy format"""
    if not isinstance(date_str, str):
        return ""
    
    formats = ["%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d"]
    for fmt in formats:
        try:
            dt_obj = datetime.datetime.strptime(date_str.strip(), fmt)
            return dt_obj.strftime("%d/%m/%Y")
        except ValueError:
            continue
    
    match = re.match(r'(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})', date_str, re.IGNORECASE)
    if match:
        day, month, year = match.groups()
        return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
    
    return ""

def process_transcript_dataframe(df):
    """Process DataFrame to ensure SBD has correct format"""
    if df.empty:
        return df
    
    df['Sbd'] = df['Sbd'].astype(str).apply(clean_sbd)
    
    invalid_sbd = df[~df['Sbd'].apply(lambda x: len(x) == 5 and x.isdigit())]
    if not invalid_sbd.empty:
        logger.warning(f"Found {len(invalid_sbd)} invalid SBD entries")
    
    df = df[df['Sbd'].apply(lambda x: len(x) == 5 and x.isdigit())]
    
    return df

def compress_image(image_bytes, max_size_mb=5):
    """Compress image if too large"""
    try:
        if len(image_bytes) <= max_size_mb * 1024 * 1024:
            return image_bytes
        
        image = Image.open(io.BytesIO(image_bytes))
        if image.width > 2000 or image.height > 2000:
            ratio = min(2000/image.width, 2000/image.height)
            new_size = (int(image.width * ratio), int(image.height * ratio))
            image = image.resize(new_size, Image.Resampling.LANCZOS)
        
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        output = io.BytesIO()
        image.save(output, format='JPEG', quality=85, optimize=True)
        return output.getvalue()
        
    except Exception as e:
        logger.warning(f"Failed to compress image: {e}")
        return image_bytes

def store_image(image_bytes, filename, session_id):
    """Store image and return path"""
    try:
        storage_path = f"ocr_sessions/{session_id}/"
        full_dir = os.path.join(settings.MEDIA_ROOT, storage_path)
        os.makedirs(full_dir, exist_ok=True)
        
        clean_filename = re.sub(r'[^\w\s.-]', '', filename.replace('/', '_').replace('\\', '_'))
        clean_filename = re.sub(r'\s+', '_', clean_filename)
        file_path = os.path.join(storage_path, clean_filename)
        
        content_file = ContentFile(image_bytes)
        saved_path = default_storage.save(file_path, content_file)
        
        return saved_path
        
    except Exception as e:
        logger.error(f"Error storing image {filename}: {e}")
        return None

def process_single_image_with_results(image_bytes, filename, api_key, processing_type, session_id, index):
    """Process single image and return detailed results"""
    try:
        image_bytes = compress_image(image_bytes, max_size_mb=3)
        image_path = store_image(image_bytes, filename, session_id)
        
        llm = ChatGoogleGenerativeAI(
            model="gemini-2.0-flash",
            temperature=0,
            google_api_key=api_key,
            request_timeout=30,
            max_retries=3
        )
        
        prompt = TRANSCRIPT_PROMPT if processing_type == "transcript" else CERTIFICATE_PROMPT
        
        image_b64 = base64.b64encode(image_bytes).decode("utf-8")
        data_uri = f"data:image/jpeg;base64,{image_b64}"
        
        message = HumanMessage(content=[
            {"type": "text", "text": prompt},
            {"type": "image_url", "image_url": {"url": data_uri}}
        ])
        
        last_error = None
        for attempt in range(2):
            try:
                response = llm.invoke([message])
                content = response.content.strip().replace('```json', '').replace('```', '')
                data = json.loads(content)
                
                if processing_type == "transcript":
                    items = []
                    for item in data.get("items", []):
                        sbd = clean_sbd(item.get("Sbd", ""))
                        thi = item.get("Thi", 0)
                        if sbd and len(sbd) == 5:
                            items.append({"Sbd": sbd, "Thi": float(thi) if thi else 0.0})
                else:
                    items = []
                    for item in data.get("items", []):
                        date_str = clean_date_string(item.get("Date_birth_VN", ""))
                        if date_str:
                            items.append({
                                "Bang_cap": item.get("Bang_cap", "").strip(),
                                "Nganh": item.get("Nganh", "").strip(),
                                "Noi_cap": item.get("Noi_cap", "").strip(),
                                "Ho_ten": item.get("Ho_ten", "").strip(),
                                "Date_birth_VN": date_str
                            })
                
                if len(items) > 0:
                    logger.info(f"✓ {os.path.basename(filename)}: {len(items)} items")
                    return {
                        "success": True,
                        "data": items,
                        "filename": filename,
                        "image_path": image_path
                    }
                else:
                    return {
                        "success": False,
                        "data": [],
                        "filename": filename,
                        "error": "Không trích xuất được dữ liệu"
                    }
                
            except json.JSONDecodeError as e:
                last_error = f"Lỗi JSON: {str(e)}"
            except Exception as e:
                last_error = f"Lỗi API: {str(e)}"
                if attempt < 1:
                    time.sleep(1)
        
        return {
            "success": False, 
            "data": [], 
            "filename": filename, 
            "error": last_error or "Thất bại"
        }
        
    except Exception as e:
        logger.error(f"Error processing {filename}: {e}")
        return {
            "success": False, 
            "data": [], 
            "filename": filename, 
            "error": str(e)
        }

def process_zip_file(zip_path, api_key, processing_type, session_id, max_images=50):
    """Process ZIP file with images"""
    all_data = []
    processed_count = 0
    image_results = []
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            image_files = []
            supported_formats = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp')
            
            for entry in zf.infolist():
                if (entry.filename.lower().endswith(supported_formats)
                    and not entry.filename.startswith('__MACOSX/')
                    and not entry.filename.startswith('.')
                    and not '/.' in entry.filename):
                    image_files.append(entry)
            
            image_files.sort(key=lambda x: x.file_size)
            
            if len(image_files) > max_images:
                logger.warning(f"Limiting to {max_images} images")
                image_files = image_files[:max_images]
            
            total_images = len(image_files)
            if total_images == 0:
                return [], []
            
            logger.info(f"Processing {total_images} images")
            update_progress(session_id, 0, total_images, f"Xử lý {total_images} ảnh...")
            
            max_workers = min(3, total_images)
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_filename = {}
                
                for i, entry in enumerate(image_files):
                    try:
                        if entry.file_size > 15 * 1024 * 1024:
                            image_results.append({
                                'filename': entry.filename,
                                'success': False,
                                'error': 'File quá lớn',
                                'data_count': 0
                            })
                            continue
                            
                        with zf.open(entry.filename) as file:
                            image_bytes = file.read()
                            if len(image_bytes) > 0:
                                future = executor.submit(
                                    process_single_image_with_results,
                                    image_bytes, entry.filename, api_key,
                                    processing_type, session_id, i
                                )
                                future_to_filename[future] = entry.filename
                                
                    except Exception as e:
                        logger.error(f"Error reading {entry.filename}: {e}")
                        image_results.append({
                            'filename': entry.filename,
                            'success': False,
                            'error': str(e),
                            'data_count': 0
                        })
                
                for future in as_completed(future_to_filename.keys(), timeout=total_images * 30):
                    filename = future_to_filename[future]
                    try:
                        result = future.result(timeout=30)
                        processed_count += 1
                        
                        image_result = {
                            'filename': os.path.basename(filename),
                            'success': result["success"],
                            'data_count': len(result["data"]) if result["success"] else 0,
                            'error': result.get("error") if not result["success"] else None
                        }
                        image_results.append(image_result)
                        
                        if result["success"]:
                            all_data.extend(result["data"])
                            update_progress(session_id, processed_count, total_images, 
                                          f"✓ {os.path.basename(filename)}")
                        else:
                            update_progress(session_id, processed_count, total_images, 
                                          f"✗ {os.path.basename(filename)}")
                        
                        time.sleep(0.2)
                        
                    except Exception as e:
                        logger.error(f"Error: {e}")
                        processed_count += 1
                        image_results.append({
                            'filename': os.path.basename(filename),
                            'success': False,
                            'error': str(e),
                            'data_count': 0
                        })
            
            success_count = sum(1 for r in image_results if r['success'])
            logger.info(f"Completed: {success_count}/{len(image_results)} success")
            update_progress(session_id, total_images, total_images, 
                          f"Xong! {success_count}/{len(image_results)} ảnh")
            
            return all_data, image_results
            
    except Exception as e:
        logger.error(f"Error processing ZIP: {e}")
        return [], []

@csrf_exempt
def upload_file(request):
    """Handle file upload"""
    if request.method == 'POST':
        try:
            form = UploadZipForm(request.POST, request.FILES)
            if form.is_valid():
                zip_file = form.cleaned_data['zip_file']
                processing_type = form.cleaned_data['processing_type']
                excel_filename = form.cleaned_data['excel_filename']
                session_id = str(uuid.uuid4())
                
                temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp')
                os.makedirs(temp_dir, exist_ok=True)
                temp_file_path = os.path.join(temp_dir, f"{session_id}_{zip_file.name}")
                
                with open(temp_file_path, 'wb') as f:
                    for chunk in zip_file.chunks():
                        f.write(chunk)
                
                from .tasks import process_images_task
                task = process_images_task.delay(session_id, temp_file_path, processing_type, excel_filename)
                
                request.session['session_id'] = session_id
                request.session['processing_type'] = processing_type
                request.session['excel_filename'] = excel_filename
                
                return JsonResponse({'success': True, 'session_id': session_id, 'task_id': task.id})
                
            return JsonResponse({'success': False, 'error': 'Form không hợp lệ'}, status=400)
            
        except Exception as e:
            logger.error(f"Upload error: {e}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    form = UploadZipForm()
    return render(request, 'apps/upload.html', {
        'form': form,
        'error_message': None
        })

@csrf_exempt
def get_progress_status(request):
    """Get progress"""
    session_id = request.GET.get('session_id')
    if not session_id:
        return JsonResponse({'error': 'Thiếu session_id'}, status=400)
    
    progress_data = get_progress(session_id)
    return JsonResponse(progress_data)

def result_page(request):
    """Display results - FIXED VERSION"""
    logger.info("=== RESULT PAGE CALLED ===")
    
    session_id = request.session.get('session_id')
    logger.info(f"Session ID: {session_id}")
    
    if not session_id:
        logger.warning("No session ID")
        return render(request, 'apps/results.html', {
            'has_data': False,
            'processing_type': 'transcript',
            'df_rows': [],
            'df_columns': [],
            'processed_images': [],
            'image_results': [],
            'error_image_filenames': [],
            'session_id': '',
            'error_message': 'Không tìm thấy session ID'
        })
    
    try:
        result_data = redis_client.get(f"result:{session_id}")
        logger.info(f"Redis data found: {result_data is not None}")
        
        if not result_data:
            logger.warning("No result in Redis")
            return render(request, 'apps/results.html', {
                'has_data': False,
                'processing_type': 'transcript',
                'df_rows': [],
                'df_columns': [],
                'processed_images': [],
                'image_results': [],
                'error_image_filenames': [],
                'session_id': session_id,
                'error_message': 'Đang xử lý... Vui lòng đợi và tải lại trang'
            })
        
        result = json.loads(result_data)
        logger.info(f"Success: {result.get('success')}, Data count: {len(result.get('data', []))}")
        
        if result.get('success') and result.get('data'):
            data = result['data']
            image_results = result.get('image_results', [])
            
            # Save to session
            request.session['extracted_data'] = data
            request.session['processing_type'] = result.get('processing_type', 'transcript')
            request.session['excel_filename'] = result.get('excel_filename', 'ocr_ketqua.xlsx')
            
            # Get processed images
            processed_images = []
            error_image_filenames = []
            
            try:
                storage_path = os.path.join(settings.MEDIA_ROOT, f"ocr_sessions/{session_id}/")
                if os.path.exists(storage_path):
                    for filename in os.listdir(storage_path):
                        if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp')):
                            img_result = next((r for r in image_results if r['filename'] == filename), None)
                            
                            processed_images.append({
                                'filename': filename,
                                'url': f"/media/ocr_sessions/{session_id}/{filename}",
                                'success': img_result['success'] if img_result else False,
                                'data_count': img_result['data_count'] if img_result else 0,
                                'error': img_result['error'] if img_result and not img_result['success'] else None
                            })
                            
                            if img_result and not img_result['success']:
                                error_image_filenames.append(filename)
                    
                    processed_images.sort(key=lambda x: x['filename'])
            except Exception as img_error:
                logger.warning(f"Error loading images: {img_error}")
            
            columns = list(data[0].keys()) if data else []
            
            logger.info(f"✓ Rendering: {len(data)} rows, {len(processed_images)} images")
            
            return render(request, 'apps/results.html', {
                'has_data': True,
                'df_rows': data,
                'df_columns': columns,
                'processing_type': result.get('processing_type', 'transcript'),
                'processed_images': processed_images,
                'image_results': image_results,
                'session_id': session_id,
                'error_image_filenames': error_image_filenames,
                'error_message': None
            })
        else:
            error_msg = result.get('error', 'Không thể xử lý')
            logger.warning(f"Processing failed: {error_msg}")
            return render(request, 'apps/results.html', {
                'has_data': False,
                'processing_type': result.get('processing_type', 'transcript'),
                'df_rows': [],
                'df_columns': [],
                'processed_images': [],
                'image_results': [],
                'error_image_filenames': [],
                'session_id': session_id,
                'error_message': error_msg
            })
            
    except Exception as e:
        logger.error(f"Error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return render(request, 'apps/results.html', {
            'has_data': False,
            'processing_type': 'transcript',
            'df_rows': [],
            'df_columns': [],
            'processed_images': [],
            'image_results': [],
            'error_image_filenames': [],
            'session_id': session_id if session_id else '',
            'error_message': f'Lỗi: {str(e)}'
        })

@csrf_exempt
def edit_record(request):
    """Handle editing"""
    if request.method == 'POST':
        try:
            changes = json.loads(request.body)
            results = []
            extracted_data = request.session.get('extracted_data', [])
            
            for change in changes:
                row_index = change.get('row_index', -1)
                field_name = change.get('field_name', '')
                new_value = change.get('new_value', '')
                
                if 0 <= row_index < len(extracted_data):
                    if field_name == 'Sbd':
                        cleaned_sbd = clean_sbd(new_value)
                        if len(cleaned_sbd) == 5 and cleaned_sbd.isdigit():
                            extracted_data[row_index][field_name] = cleaned_sbd
                            results.append({'success': True, 'row_index': row_index})
                        else:
                            results.append({'success': False, 'error': 'SBD không hợp lệ', 'row_index': row_index})
                    elif field_name == 'Thi':
                        try:
                            score = float(new_value)
                            if 0 <= score <= 10:
                                extracted_data[row_index][field_name] = score
                                results.append({'success': True, 'row_index': row_index})
                            else:
                                results.append({'success': False, 'error': 'Điểm 0-10', 'row_index': row_index})
                        except ValueError:
                            results.append({'success': False, 'error': 'Phải là số', 'row_index': row_index})
                    elif field_name == 'Date_birth_VN':
                        cleaned_date = clean_date_string(new_value)
                        if cleaned_date:
                            extracted_data[row_index][field_name] = cleaned_date
                            results.append({'success': True, 'row_index': row_index})
                        else:
                            results.append({'success': False, 'error': 'Ngày không hợp lệ', 'row_index': row_index})
                    else:
                        extracted_data[row_index][field_name] = str(new_value).strip()
                        results.append({'success': True, 'row_index': row_index})
                else:
                    results.append({'success': False, 'error': 'Không tồn tại', 'row_index': row_index})
            
            request.session['extracted_data'] = extracted_data
            request.session.modified = True
            
            return JsonResponse(results, safe=False)
            
        except Exception as e:
            logger.error(f"Error: {e}")
            return JsonResponse([{'success': False, 'error': str(e)}], safe=False)
    
    return JsonResponse([{'success': False, 'error': 'Invalid'}], safe=False)

@csrf_exempt  
def replace_image(request):
    """Replace image"""
    if request.method == 'POST':
        try:
            image_file = request.FILES.get('image')
            row_index = int(request.POST.get('row_index', -1))
            session_id = request.session.get('session_id')
            
            if not all([image_file, session_id, row_index >= 0]):
                return JsonResponse({'success': False, 'error': 'Thiếu dữ liệu'})
            
            if not image_file.content_type.startswith('image/'):
                return JsonResponse({'success': False, 'error': 'Không phải ảnh'})
            
            if image_file.size > 10 * 1024 * 1024:
                return JsonResponse({'success': False, 'error': 'Quá lớn'})
            
            image_bytes = image_file.read()
            filename = f"replaced_{row_index}_{image_file.name}"
            image_path = store_image(image_bytes, filename, session_id)
            
            if image_path:
                return JsonResponse({
                    'success': True,
                    'new_path': f"/media/{image_path}",
                    'message': 'Thành công'
                })
            else:
                return JsonResponse({'success': False, 'error': 'Không lưu được'})
                
        except Exception as e:
            logger.error(f"Error: {e}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid'})

def download_excel(request):
    """Download Excel"""
    extracted_data = request.session.get('extracted_data')
    excel_filename = request.session.get('excel_filename', 'ocr_ketqua.xlsx')
    processing_type = request.session.get('processing_type', 'transcript')
    
    if not extracted_data:
        return redirect('upload_file')

    try:
        if isinstance(extracted_data, str):
            extracted_data = json.loads(extracted_data)
        
        df = pd.DataFrame(extracted_data)
        
        if df.empty:
            return JsonResponse({'error': 'No data'}, status=400)
        
        if processing_type == "transcript":
            df = process_transcript_dataframe(df)
            if 'Thi' in df.columns:
                df['Thi'] = pd.to_numeric(df['Thi'], errors='coerce').fillna(0.0)

            output_cols = [
                'TT', 'Lớp môn nhập AAIS', 'Tài khoản', 'SBD', 'Lớp', 'MSSV', 
                'Họ và tên', 'Ngày sinh', 'X', 'QT', 'KT', 'Thi', 'Điểm học phần', 
                'Thang điểm chữ', 'Thang điểm 4', 'Ghi chú'
            ]
            
            rows = []
            for idx, row in df.iterrows():
                new_row = {col: None for col in output_cols}
                new_row['TT'] = idx + 1
                new_row['SBD'] = row.get('Sbd', '')
                new_row['Thi'] = row.get('Thi', 0.0)
                rows.append(new_row)
                
            df_final = pd.DataFrame(rows, columns=output_cols)

        elif processing_type == "certificate":
            df_final = pd.DataFrame([{
                "TT": i + 1,
                "Đơn vị đào tạo": "Viện ĐT&PT học tập suốt đời",
                "Đơn vị liên kết": "",
                "Ngành học HOU": "",
                "Lớp": "",
                "Họ và tên": d.get("Ho_ten", ""),
                "Ngày sinh": d.get("Date_birth_VN", ""),
                "Mã SV": "",
                "Văn Bằng": d.get("Bang_cap", ""),
                "Ngành": d.get("Nganh", ""),
                "Nơi cấp": d.get("Noi_cap", ""),
                "TT_2": "",
                "Tên Học Phần": "",
                "Mã môn": "",
                "Số TC": "",
                "Tài khoản học": ""
            } for i, d in enumerate(extracted_data)])

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            if processing_type == "transcript":
                df_final.to_excel(writer, index=False, sheet_name='Sheet1')
                ws = writer.sheets['Sheet1']
                
                if 'SBD' in output_cols:
                    sbd_col_idx = output_cols.index("SBD") + 1
                    for i in range(2, ws.max_row + 1):
                        sbd_cell = ws[f'{get_column_letter(sbd_col_idx)}{i}']
                        sbd_cell.number_format = '@'
                
                if 'Thi' in output_cols:
                    thi_col_idx = output_cols.index("Thi") + 1
                    for i in range(2, ws.max_row + 1):
                        thi_cell = ws[f'{get_column_letter(thi_col_idx)}{i}']
                        thi_cell.number_format = '0.0'
            else:
                df_final.to_excel(writer, index=False, startrow=9, sheet_name='Sheet1')
                
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.read(), 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{excel_filename}"'
        
        buffer.close()
        return response
        
    except Exception as e:
        logger.error(f"Error in download_excel: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({'error': f'Lỗi: {str(e)}'}, status=500)